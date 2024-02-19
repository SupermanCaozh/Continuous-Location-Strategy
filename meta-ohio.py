import numpy as np
import pandas as pd
import math
import matplotlib.pyplot as plt
import os
from openpyxl import load_workbook

plt.rcParams["font.family"] = ['SimHei']  # 用来设定字体样式
plt.rcParams['font.sans-serif'] = ['SimHei']  # 用来设定无衬线字体样式

region_pop = [4100, 7800, 8100, 10700, 11500, 9300, 10100, 8800, 5300, 5100, 7700, 4300,
              6200, 8700, 10500, 12800, 13900, 14900, 12600, 13700, 16700, 17400, 9200, 6700,
              7200, 9400, 15600, 13800, 14500, 13700, 16700, 15200, 13800, 10300, 7500, 5800,
              10300, 11800, 10500, 15600, 13700, 10200, 15800, 14100, 11900, 9800, 8500, 6800,
              200, 100, 200, 400, 600, 1200, 12400, 10800, 13500, 10300, 7800, 5400,
              2600, 17200, 18600, 15500, 9900, 7100,
              500, 12000, 11700, 8700, 6400]
k = 2.5  # 度量因子,观察地图可得
x_col = np.arange(1, 13)
y_row = np.arange(1, 8)  # 在该问题表示中,图中左下为(0.5,0.5)
regions_cord = []
for i in y_row[:-2]:
    for j in x_col:
        regions_cord.append([j, i])  # [col,row]
for i in x_col[6:]:
    regions_cord.append([i, 6])
for i in x_col[7:]:
    regions_cord.append([i, 7])
rate_trl = 0.12
num_cluster = 6


def cal_dis(a, b):
    """
    计算两坐标点欧式距离
    :param a: [xa,ya]
    :param b: [xb,yb]
    :return:
    """
    distance = np.sqrt(np.sum(np.square(np.array(a) - np.array(b))))
    return distance


def cost2(clusters, num):
    """
    计算经营成本和出行成本
    :param clusters: [gravity,region1,...,regionM]
    :return:
    """
    # 计算出行成本
    cost_trl = []
    # 计算经营成本
    cost_opr = []
    for i in range(num):
        cost1 = 0
        cost22 = 1500
        cost3 = 4  # 计算人事费
        pops = 0
        gravity = clusters[i][0]
        for cord in clusters[i][1:]:
            d = cal_dis(gravity, cord) * k
            idx = regions_cord.index(cord)
            pop = region_pop[idx]
            cost1 += pop * rate_trl * d

            pops += pop
        cost_trl.append(cost1)
        # 计算经营成本
        cost22 += max(0, 500 * (np.floor(pops / 100000) - 1))
        cost22 = cost22 * (22 + 4)

        cost3 += max(0, (np.floor(pops / 100000) - 1) * 1)
        cost3 = cost3 * 21000
        cost_opr.append(cost22 + cost3)
    return np.sum(cost_opr), np.sum(cost_trl)


def optimizer(num):
    popsize = 50
    Max_iter = 500
    lbs = [0.5, 0.5]
    ubs = [12.5, 7.5]
    cost_opr_convergence = []
    cost_trl_convergence = []
    cost_convergence = []
    alpha_pos = None
    beta_pos = None
    delta_pos = None
    alpha_cost = float(math.inf)
    beta_cost = float(math.inf)
    delta_cost = float(math.inf)
    best_cost_split = np.zeros(2)
    best_clusters = []

    # initialize the location
    pop_gravity = []
    for i in range(popsize):
        gravities_x = np.random.uniform(0.5, 12.5, num)
        gravities_y = np.random.uniform(0.5, 7.5, num)
        gravities = np.zeros((num, 2))
        gravities[:, 0] = gravities_x
        gravities[:, 1] = gravities_y
        pop_gravity.append(gravities)

    for iter in range(Max_iter):
        print("Iteration {} starts. The temp best layout: {}".format(iter + 1, alpha_pos))

        # 对网点具体位置从西南向东北排序，即按照距离原点的距离
        for i in range(popsize):
            g = pop_gravity[i]
            g_sort = sorted(g, key=lambda x: (x[0] ** 2 + x[1] ** 2) ** (1 / 2))
            g_sort = np.array(g_sort)
            pop_gravity[i] = g_sort

        for i in range(popsize):
            gravities = pop_gravity[i]
            clusters = [[g] for g in gravities]
            # 将各方格区域分配给网点
            for cord in regions_cord:
                ds = [cal_dis(cord, gravities[w]) for w in range(num)]
                d = min(ds)
                idx = ds.index(d)
                clusters[idx].append(cord)

            # 评价当前布局的总成本
            cost_opr, cost_trl = cost2(clusters, num)
            cost = cost_opr + cost_trl

            if cost < alpha_cost:
                alpha_pos = gravities
                alpha_cost = cost
                best_cost_split[0], best_cost_split[1] = cost_opr, cost_trl
                best_clusters = clusters

            elif cost < beta_cost:
                beta_pos = gravities
                beta_cost = cost

            elif cost < delta_cost:
                delta_pos = gravities
                delta_cost = cost

        # 更新重心
        # std = 1 - (iter + 1) / Max_iter
        std = np.exp(-100 * (iter + 1) / Max_iter)
        error = np.random.normal(0, std, np.shape(alpha_pos))
        prey_pos = 0.5 * alpha_pos + 0.3 * beta_pos + 0.2 * delta_pos + error

        for i in range(popsize):
            new_pos = np.zeros(np.shape(alpha_pos))
            pos = pop_gravity[i]
            for j in range(num):
                r = np.random.uniform(-2, 2, 2)
                new_pos[j] = prey_pos[j] - r * abs(prey_pos[j] - pos[j])
                for z in range(2):
                    if new_pos[j][z] < lbs[z]:
                        u = np.random.uniform(0, 1)
                        new_pos[j][z] = pos[j][z] + u * (lbs[z] - pos[j][z])
                    if new_pos[j][z] > ubs[z]:
                        u = np.random.uniform(0, 1)
                        new_pos[j][z] = pos[j][z] + u * (ubs[z] - pos[j][z])
            pop_gravity[i] = new_pos

        # 记录最优成本和布局
        cost_convergence.append(alpha_cost)
        cost_opr_convergence.append(best_cost_split[0])
        cost_trl_convergence.append(best_cost_split[1])

        if (iter + 1) % 500 == 0:
            fig = plt.figure(dpi=200)
            plt.plot(cost_convergence, linewidth=2)
            plt.grid(ls="--")
            plt.xlabel("迭代次数")
            plt.ylabel("适应度值")
            plt.tight_layout()
            plt.show()

    return alpha_pos, alpha_cost, best_cost_split, best_clusters, cost_convergence, cost_opr_convergence, cost_trl_convergence


def record(target_dir, file_name, sheet_name, data, col_names):
    if not os.path.exists(target_dir):
        os.makedirs(target_dir)
    file = os.path.join(target_dir, file_name + ".xlsx")
    if not os.path.exists(file):
        df = pd.DataFrame(columns=col_names)
        df.to_excel(file, sheet_name=sheet_name,
                    index=False)

    book = load_workbook(file)
    # df = pd.DataFrame(data)
    df = data

    if sheet_name in book.sheetnames:
        df1 = pd.DataFrame(pd.read_excel(file, sheet_name=sheet_name))
        writer = pd.ExcelWriter(file, engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        df_rows = df1.shape[0]
        df.to_excel(writer, sheet_name=sheet_name, startrow=df_rows + 1, index=False,
                    header=False)
        writer.save()
        writer.close()

    else:  # need to create a new sheet to record
        writer = pd.ExcelWriter(file, engine='openpyxl')
        writer.book = book
        df.to_excel(writer, sheet_name=sheet_name, header=True, index=None)
        writer.save()
        writer.close()


# best_layout, min_cost, (
#     min_cost_opr, min_cost_trl), cost_convergence, cost_opr_convergence, cost_trl_convergence = optimizer()

target_dir = r"C:\Users\caoze\Downloads\研究生课件\供应链"
file_name = "data"

# 讨论开设不同数量的管理局时,总成本及其构成情况
# nums_cluster = np.arange(1, 11)
nums_cluster = [10]
# num_runs = 5
num_runs = 1

for num in nums_cluster:
    col_names = ["nodes", "minimum total cost", "minimum operation cost", "minimum travelling cost", "region clusters"]
    min_costs = pd.DataFrame(index=np.arange(1, num_runs + 1), columns=col_names)
    for run in range(num_runs):
        best_layout, min_cost, (min_cost_opr, min_cost_trl), best_clusters, \
        cost_convergence, cost_opr_convergence, cost_trl_convergence = optimizer(num)

        min_costs.loc[run + 1, "nodes"] = str(best_layout.tolist())  # 读取的时候eval回来
        min_costs.loc[run + 1, "minimum total cost"] = min_cost
        min_costs.loc[run + 1, "minimum operation cost"] = min_cost_opr
        min_costs.loc[run + 1, "minimum travelling cost"] = min_cost_trl
        min_costs.loc[run + 1, "region clusters"] = str(best_clusters)

    sheet_name = "cluster {}".format(num)

    # record(target_dir, file_name, sheet_name, min_costs, col_names)

# 针对总成本最小的群落个数,分析其各项成本构成
# # 记录各管理局的经营成本&记录各方格区域居民的出行成本
# cost_opr = np.zeros((num, 3))  # 0列记录租金费用,1列记录人事费用,2列记录该网点总经营成本
# cost_trl = np.zeros((12, 7))
#
# for i in range(num):
#     cost1 = 0
#     cost22 = 1500
#     cost3 = 4  # 计算人事费
#     pops = 0
#     gravity = best_clusters[i][0]
#     for cord in best_clusters[i][1:]:
#         d = cal_dis(gravity, cord) * k
#         idx = regions_cord.index(cord)
#         pop = region_pop[idx]
#         cost1 = pop * rate_trl * d
#         cost_trl[cord[0] - 1, cord[1] - 1] = cost1  # 在表格中记录当前网格区域内居民出行成本
#
#         pops += pop
#     # 计算经营成本
#     cost22 += 500 * (np.floor(pops / 100000) - 1)
#     cost22 = cost22 * (22 + 4)
#     cost3 += (np.floor(pops / 100000) - 1) * 1
#     cost3 = cost3 * 21000
#     cost4 = cost22 + cost3
#
#     cost_opr[i, 0] = cost22
#     cost_opr[i, 1] = cost3
#     cost_opr[i, 2] = cost4
